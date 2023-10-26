# 파이썬에서 openpyxl 라이브러리를 사용해서 전자제품 판매 리스트를 작성해줘. id, name, 수량,가격이 있는 판매 리스트를 100개의 행으로생성해서 c:\work\sales.xlsx 파일로 저장해줘

import openpyxl

# Create a new workbook and select the default sheet
workbook = openpyxl.Workbook()
sheet = workbook.active

# Set the column headers
sheet['A1'] = "ID"
sheet['B1'] = "Name"
sheet['C1'] = "Quantity"
sheet['D1'] = "Price"

# Generate and insert data for 100 rows
for row in range(2, 102):  # Rows 2 to 101 (100 rows)
    sheet.cell(row=row, column=1, value=row - 1)  # ID
    sheet.cell(row=row, column=2, value=f"Product {row - 1}")  # Name
    sheet.cell(row=row, column=3, value=1)  # Quantity (you can modify this as needed)
    sheet.cell(row=row, column=4, value=100.0 + (row - 1))  # Price (you can modify this as needed)

# Save the workbook to a file
file_path = r'c:\work\sales.xlsx'
workbook.save(file_path)

print(f"Sales list saved to {file_path}")
